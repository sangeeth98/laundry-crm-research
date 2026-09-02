"""
Export 51 B2B Laundry CRM & Operations Platforms to a formatted multi-sheet Excel workbook.
Reads the single verified source of truth: data/laundry_crm_51_companies_verified.json.
"""

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_FILE = Path("data/laundry_crm_51_companies_verified.json")
EXCEL_OUTPUT = Path("data/Laundry_CRM_Market_Intelligence_Master.xlsx")

def style_header_row(ws, cols):
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    for col_idx in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

def auto_fit_columns(ws):
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            cell.border = thin_border
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 48)

def main():
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        comps = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # -------------------------------------------------------------
    # Sheet 1: Master Platforms Matrix (All 51)
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="All 51 CRM Platforms")
    headers1 = [
        "#", "Platform Name", "Legal Entity", "Tier", "Market Status",
        "Business Model", "HQ Origin", "Countries Count", "Inception",
        "Verified Actual Revenue", "Filing Authority", "Projected Run-Rate",
        "Headcount", "Starter Price ($/mo)", "Website"
    ]
    ws1.append(headers1)
    style_header_row(ws1, headers1)

    for idx, c in enumerate(comps, 1):
        act = c.get("actual_revenue", {})
        ws1.append([
            idx,
            c.get("name", ""),
            c.get("legal_entity", ""),
            c.get("tier_label", "").split(":")[0],
            c.get("status_category", "").upper(),
            c.get("business_model_category", "").replace("_", " ").title(),
            c.get("market", {}).get("countries_list", [""])[0],
            c.get("market", {}).get("country_count", 1),
            c.get("start_date", ""),
            act.get("reported_figure") if act.get("actual_status") == "verified" else "Undisclosed (Private)",
            act.get("source_authority", "N/A"),
            act.get("projected_figure", "N/A"),
            c.get("employee_count", {}).get("current", 0),
            c.get("starter_price_usd", 0.0),
            c.get("site_links", {}).get("active", [""])[0]
        ])
    auto_fit_columns(ws1)

    # -------------------------------------------------------------
    # Sheet 2: Founders & Leadership Intelligence
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Founders & Pedigree")
    headers2 = [
        "#", "Platform Name", "Key Founders & Executives",
        "Educational Pedigree & Alma Mater", "Career Trajectory",
        "Operating Principles & Philosophy", "Current Expansion Focus"
    ]
    ws2.append(headers2)
    style_header_row(ws2, headers2)

    for idx, c in enumerate(comps, 1):
        fh = c.get("founder_history", {})
        ws2.append([
            idx,
            c.get("name", ""),
            "; ".join(c.get("founders", [])),
            fh.get("pedigree_education", "N/A"),
            fh.get("career_trajectory", "N/A"),
            fh.get("life_operating_principles", "N/A"),
            fh.get("latest_strategic_focus", "N/A")
        ])
    auto_fit_columns(ws2)

    # -------------------------------------------------------------
    # Sheet 3: Financial Disclosures & Estimation Methodology
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Financials & Grounded Sources")
    headers3 = [
        "#", "Platform Name", "Tier", "Actual Status",
        "Verified Public Figure", "Filing Period", "Registry Authority",
        "Source Citation / Link", "Projected Run-Rate", "Normalized ARR ($M)",
        "Projection Methodology & Footprint Calculation"
    ]
    ws3.append(headers3)
    style_header_row(ws3, headers3)

    for idx, c in enumerate(comps, 1):
        act = c.get("actual_revenue", {})
        ws3.append([
            idx,
            c.get("name", ""),
            c.get("tier_label", "").split(":")[0],
            "Verified Disclosure" if act.get("actual_status") == "verified" else "Undisclosed (Private)",
            act.get("reported_figure", "Not Publicly Disclosed"),
            act.get("period", "N/A"),
            act.get("source_authority", "Private Entity"),
            act.get("source_citation", act.get("registry_band", "Confidential / Paid Registry")),
            act.get("projected_figure", "N/A"),
            c.get("est_revenue_usd_m", 0.0),
            act.get("projection_methodology", "Analytical Model")
        ])
    auto_fit_columns(ws3)

    # -------------------------------------------------------------
    # Sheet 4: Strategic Moat & CRM Blueprints
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Strategic Playbooks & Lessons")
    headers4 = [
        "#", "Platform Name", "Tier", "Market Status",
        "Success Flywheel or Failure Post-Mortem",
        "Competitor Vulnerabilities & Actionable CRM Takeaways"
    ]
    ws4.append(headers4)
    style_header_row(ws4, headers4)

    for idx, c in enumerate(comps, 1):
        ss = c.get("strategic_story", {})
        ws4.append([
            idx,
            c.get("name", ""),
            c.get("tier_label", "").split(":")[0],
            c.get("status_category", "").upper(),
            ss.get("success_or_failure_analysis", "N/A"),
            ss.get("vulnerabilities_lessons", "N/A")
        ])
    auto_fit_columns(ws4)

    wb.save(EXCEL_OUTPUT)
    print(f"Successfully exported {len(comps)} companies to {EXCEL_OUTPUT}")

if __name__ == "__main__":
    main()
